import tkinter as tk
from tkinter import ttk
import os
from openpyxl import Workbook, load_workbook

def login():
    username = ""
    password = ""
    if username_entry.get() == username and password_entry.get() == password:
        for widget in window.winfo_children():
            widget.destroy()
        newwindow()
    else:
        error_message.set("Invalid login. Please try again.")

def savedata():
    name = name_entry.get()
    address = address_entry.get()
    gender = gender_var.get()
    email = email_entry.get()
    phone = phone_entry.get()
    item = item_var.get()
    quantity = quantity_entry.get()
    amount = amount_entry.get()

    savexl = "data.xlsx"

    try:
        if savexl in [f for f in os.listdir() if os.path.isfile(f)]:
            wb = load_workbook(savexl)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        ws.append([name, address, gender, email, phone, item, quantity, amount])
        wb.save(savexl)
        error_message.set("Data saved successfully.")
    except Exception as e:
        error_message.set("Error")
    clear_entry()

def clear_entry():
    name_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)
    gender_var.set("Male")
    email_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    item_var.set("Rice")
    quantity_entry.delete(0, tk.END)
    amount_entry.delete(0, tk.END)

def newwindow():
    global name_entry, address_entry, gender_var, email_entry, phone_entry, item_var, quantity_entry, amount_entry

    frame = tk.Frame(window, bg='#5fc5dc')
    frame.pack(expand=True)

    name_label = tk.Label(frame, text="Name", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
    name_entry = tk.Entry(frame, font=("Arial", 12))

    address_label = tk.Label(frame, text="Address", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
    address_entry = tk.Entry(frame, font=("Arial", 12))

    gender_label = tk.Label(frame, text="Gender", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
    gender_var = tk.StringVar()
    gender_option = ttk.OptionMenu(frame, gender_var, "Male", "Female", "Other")

    email_label = tk.Label(frame, text="Email", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))   
    email_entry = tk.Entry(frame, font=("Arial", 12))
    

    phone_label = tk.Label(frame, text="Phone No.", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))   
    phone_entry = tk.Entry(frame, font=("Arial", 12))
    

    item_label = tk.Label(frame, text="Item", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12)) 
    item_var = tk.StringVar()
    item_var.set("Rice")
    item_option = ttk.OptionMenu(frame, item_var, "Rice", "Sugar", "Milk")

    quantity_label = tk.Label(frame,text="Quantity", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
    quantity_entry = tk.Entry(frame,font=("Arial",12))
    
    amount_label =tk.Label(frame,text="Amount", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
    amount_entry =tk.Entry(frame,font=("Arial",12))

    save_button = tk.Button(frame, text="Save", bg="#e9692c", fg="#FFFFFF", font=("Helvetica", 12), command=savedata)
    

    name_label.grid(row=0, column=0, padx=10, pady=10)
    name_entry.grid(row=0, column=1, padx=10, pady=10)
    address_label.grid(row=1, column=0, padx=10, pady=10)
    address_entry.grid(row=1, column=1, padx=10, pady=10)
    gender_label.grid(row=2, column=0, padx=10, pady=10)
    gender_option.grid(row=2, column=1, padx=10, pady=10)
    email_label.grid(row=3, column=0, padx=10, pady=10)
    email_entry.grid(row=3, column=1, padx=10, pady=10)
    phone_label.grid(row=4, column=0, padx=10, pady=10)
    phone_entry.grid(row=4, column=1, padx=10, pady=10)
    item_label.grid(row=5, column=0, padx=10, pady=10)
    item_option.grid(row=5,column=1,padx=10,pady=10)
    quantity_label.grid(row=6,column=0,padx=10,pady=10)
    quantity_entry.grid(row=6,column=1,padx=10,pady=10)
    amount_label.grid(row=7,column=0,padx=10,pady=10)
    amount_entry.grid(row=7,column=1,padx=10,pady=10)
    save_button.grid(row=8, column=1, columnspan=2, padx=10)

window = tk.Tk()
window.title("CRM login")
window.geometry('800x800')
window.configure(bg='#5fc5dc')

frame = tk.Frame(window, bg='#5fc5dc')
login_label = tk.Label(frame, text="Login", bg='#5fc5dc', fg="#e9692c", font=("Helvetica", 25))
username_label = tk.Label(frame, text="Username", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
username_entry = tk.Entry(frame, font=("Arial", 12))
password_entry = tk.Entry(frame, show="*", font=("Arial", 12))
password_label = tk.Label(frame, text="Password", bg='#5fc5dc', fg="#FFFFFF", font=("Helvetica", 12))
login_button = tk.Button(frame, text="Login", bg="#e9692c", fg="#FFFFFF", font=("Helvetica", 12), command=login)

error_message = tk.StringVar()
error_label = tk.Label(frame, textvariable=error_message, bg='#5fc5dc', fg="#FF3333", font=("Helvetica", 12))

login_label.grid(row=0, column=0, columnspan=2, sticky="news", pady=20)
username_label.grid(row=1, column=0)
username_entry.grid(row=1, column=1, pady=10)
password_label.grid(row=2, column=0)
password_entry.grid(row=2, column=1, pady=10)
login_button.grid(row=3, column=1, columnspan=2, pady=30)
error_label.grid(row=4, column=0, columnspan=2, pady=10)

frame.pack(expand=True)

window.mainloop()